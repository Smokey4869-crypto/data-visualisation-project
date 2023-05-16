from openpyxl import load_workbook
import json

# https://www.pythonexcel.com/openpyxl-load-workbook-function.php
# https://openpyxl.readthedocs.io/en/stable/optimized.html?highlight=load_workbook#read-only-mode
####
# read_only=True 
# write_only=True
# data_only 

wb = load_workbook(filename = 'data/reasons_for_moving_draft.xlsx')
sheet = wb['Sheet1']

reason_1 = sheet['C1'].value.replace("reasons", "").strip()
reason_2 = sheet['F1'].value.replace("reasons", "").strip()
reason_3 = sheet['K1'].value.replace("reasons", "").strip()
reason_4 = sheet['Q1'].value.replace("reasons", "").strip()

y = {
    "2010-2011": [],
    "2011-2012": [],
    "2012-2013": [],
    "2013-2014": [],
    "2014-2015": [],
    "2015-2016": [],
    "2016-2017": [],
    "2017-2018": [],
    "2018-2019": [],
}

# JSON template per year
data_template = [
    {
        "nodeData": {
            "name": reason_1,
            "value": ""
        },
        "subData": [
            {
                "nodeData": {
                    "name": sheet['C2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData" : {
                    "name": sheet['D2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['E2'].value.strip('.\n'),
                    "value": ""
                }
            }
        ]
    },
    {
        "nodeData": {
            "name": reason_2,
            "value": ""
        },
        "subData": [
            {
                "nodeData": {
                    "name": sheet['F2'].value.strip('.\n'),
                    "value": ""
                }
            },
            {
                "nodeData" : {
                    "name": sheet['G2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['H2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['I2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['J2'].value.strip('.\n'),
                    "value": ""
                }
            }
        ]
    },
    {
        "nodeData": {
            "name": reason_3,
            "value": ""
        },
        "subData": [
            {
                "nodeData": {
                    "name": sheet['K2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['L2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['M2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['N2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['O2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['P2'].value.strip('.\n'),
                    "value": ""
                },
            },
        ]
    },
    {
        "nodeData": {
            "name": reason_4,
            "value": ""
        },
        "subData": [
            {
                "nodeData": {
                    "name": sheet['Q2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['R2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['S2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['T2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['U2'].value.strip('.\n'),
                    "value": ""
                },
            },
            {
                "nodeData": {
                    "name": sheet['V2'].value.strip('.\n'),
                    "value": ""
                },
            },
        ]
    },
]


ws = wb.active

# Get the index of the last non-empty row
last_row = ws.max_row

# Print the value of each cell in each column from rows 3 to last row

for row in range(3, last_row+1):
    period = ws['A' + str(row)].value
    
    for data in data_template:
        if data["nodeData"]["name"] == reason_1:
            data["nodeData"]["value"] = round(ws['W' + str(row)].value, 1)
            for i in range(ord('C'), ord('E')+1):
                if ws[chr(i)+str(row)].value == "N":
                    data["subData"][i - ord('C')]["nodeData"]["value"] = 0
                else:
                    data["subData"][i - ord('C')]["nodeData"]["value"] = round(float(ws[chr(i)+str(row)].value), 1)

        if data["nodeData"]["name"] == reason_2:
            data["nodeData"]["value"] = round(ws['X'+ str(row)].value, 1)  
            for i in range(ord('F'), ord('J')+1):
                if ws[chr(i)+str(row)].value == "N":
                    data["subData"][i - ord('F')]["nodeData"]["value"] = 0
                else:
                    data["subData"][i - ord('F')]["nodeData"]["value"] = round(float(ws[chr(i)+str(row)].value), 1)

        if data["nodeData"]["name"] == reason_3:
            data["nodeData"]["value"] = round(ws['Y'+ str(row)].value, 1) 
            for i in range(ord('K'), ord('P')+1):
                if ws[chr(i)+str(row)].value == "N":
                    data["subData"][i - ord('K')]["nodeData"]["value"] = 0
                else:
                    data["subData"][i - ord('K')]["nodeData"]["value"] = round(float(ws[chr(i)+str(row)].value), 1)
            
        if data["nodeData"]["name"] == reason_4:
            data["nodeData"]["value"] = round(ws['Z'+ str(row)].value, 1)
            for i in range(ord('Q'), ord('V')+1):
                if ws[chr(i)+str(row)].value == "N":
                    data["subData"][i - ord('Q')]["nodeData"]["value"] = 0
                else:
                    data["subData"][i - ord('Q')]["nodeData"]["value"] = round(float(ws[chr(i)+str(row)].value), 1)


        y[period].append(data)

# Serializing JSON object
json_object = json.dumps(y, indent=4)

# Write JSON to new file
with open("data/reasons_for_moving_draft.json", "w") as outfile:
    outfile.write(json_object)

