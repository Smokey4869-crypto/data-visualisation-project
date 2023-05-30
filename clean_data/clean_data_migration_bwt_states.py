import openpyxl
import json

def convert_xlsx_to_json(file_path):
    workbook = openpyxl.load_workbook(file_path)
    json_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        
        headers = [str(cell.value) for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))

        json_data[sheet_name] = data

    return json_data

# Specify the path to your XLSX file
xlsx_file_path = 'data/State_to_State.xlsx'

# Convert the XLSX file to JSON objects
json_objects = convert_xlsx_to_json(xlsx_file_path)

# Write all JSON objects to a single file
output_file_path = 'data/state_to_state.json'
with open(output_file_path, 'w') as file:
    json.dump(json_objects, file, indent=4)

print(f"All JSON objects have been written to {output_file_path}")